import json
from openpyxl import load_workbook

def load_db_struct(json_path: str) -> tuple[dict, dict]:
    """Carica la struttura del database e le chiavi primarie."""
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    return data.get("tables", {}), data.get("table_pks", {})

def flatten_db_struct(tables: dict, table_pks: dict) -> list[dict]:
    """
    Crea una lista piatta di tutte le colonne con metadati:
    table_name, column_name, human_readable_col_name, is_pk
    """
    flat_data = []
    for table_name, columns in tables.items():
        pks = set(table_pks.get(table_name, []))
        for col in columns:
            flat_data.append({
                "table_name": table_name,
                "column_name": col.get("column_name"),
                "human_readable_col_name": col.get("human_readable_col_name", ""),
                "is_pk": col.get("column_name") in pks
            })
    return flat_data

def fill_excel(excel_path: str, flat_data: list[dict], output_path: str) -> None:
    """
    Inserisce i dati nel file Excel esistente mantenendo la struttura.
    Scrive solo nelle colonne A, D, E, F.
    """
    wb = load_workbook(excel_path)
    ws = wb.active  # se serve un foglio specifico, si può parametrizzare

    start_row = 2  # supponendo riga 1 = intestazioni
    for i, row_data in enumerate(flat_data, start=start_row):
        ws[f"A{i}"] = row_data["table_name"]
        ws[f"D{i}"] = row_data["column_name"]
        ws[f"E{i}"] = row_data["human_readable_col_name"]
        ws[f"F{i}"] = "N" if row_data["is_pk"] else "S"

    wb.save(output_path)
    print(f"✅ File aggiornato salvato come: {output_path}")

def main():
    json_path = "db_repr.json"
    excel_path = "AIMP_Table.xlsx"
    output_path = "AIMP_Table_completed.xlsx"

    tables, table_pks = load_db_struct(json_path)
    flat_data = flatten_db_struct(tables, table_pks)
    fill_excel(excel_path, flat_data, output_path)

if __name__ == "__main__":
    main()
